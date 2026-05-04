from typing import Iterator

from openpyxl import load_workbook

from pipeforge.config.models import ExcelInputConfig
from pipeforge.core.registry import register_plugin
from pipeforge.plugins.base import InputPlugin


def read_excel_rows(file: str, sheet: str | None = None) -> tuple[list[str], Iterator[tuple]]:
    """读取 Excel 文件，返回 (columns, rows) 元组。"""
    wb = load_workbook(file, read_only=True, data_only=True)

    if sheet is None:
        ws = wb.active
    else:
        if sheet not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        wb.close()
        raise ValueError("Excel file has no header row (first row is empty)")

    if header is None or all(h is None for h in header):
        wb.close()
        raise ValueError("Excel file has an empty header row")

    columns = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header)]

    def row_generator():
        for row in rows_iter:
            if row is not None:
                yield row
        wb.close()

    return columns, row_generator()


@register_plugin("excel", "input")
class ExcelInputPlugin(InputPlugin):
    """从 Excel 文件读取数据并写入 SQLite。"""

    @classmethod
    def config_model(cls) -> type[ExcelInputConfig]:
        return ExcelInputConfig

    def execute(self, context, config: ExcelInputConfig) -> None:
        columns, rows = read_excel_rows(config.file, config.sheet)
        context.db.create_table(self.table_name, columns)
        with context.db.transaction():
            for row in rows:
                context.db.insert_row(self.table_name, row)
        context.logger.info(f"Input '{self.label}': loaded data into table '{self.table_name}'")
