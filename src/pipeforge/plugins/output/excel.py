import copy
import os
import re
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell

from pipeforge.config.models import ExcelOutputConfig
from pipeforge.core.registry import register_plugin
from pipeforge.plugins.base import OutputPlugin


def resolve_filename(filename_template: str | None, scene_name: str) -> str:
    """Resolve variable placeholders in the filename template."""
    if filename_template is None:
        filename_template = "{{scene_name}}_{{date:%Y%m%d}}.xlsx"

    now = datetime.now()
    result = filename_template
    result = result.replace("{{scene_name}}", scene_name)
    result = result.replace("{{timestamp}}", str(int(now.timestamp())))

    def replace_date(match):
        fmt = match.group(1)
        return now.strftime(fmt)

    result = re.sub(r"\{\{date:(.+?)\}\}", replace_date, result)
    result = re.sub(r"\{\{time:(.+?)\}\}", replace_date, result)
    return result


@register_plugin("excel", "output")
class ExcelOutputPlugin(OutputPlugin):
    """Excel template output plugin — three-phase write to preserve styles,
    column widths, and freeze panes."""

    @classmethod
    def config_model(cls) -> type[ExcelOutputConfig]:
        return ExcelOutputConfig

    def execute(self, context, config: ExcelOutputConfig) -> None:
        template_path = os.path.join(context.yaml_dir, config.template)
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template file not found: {template_path}")

        filename = resolve_filename(config.filename, context.scene_name)
        output_dir = config.output_dir or "./output/"
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, filename)
        context.output_path = output_path

        rows = context.db.query(f'SELECT * FROM "{config.source_table}"')
        source_columns = context.db.get_column_names(config.source_table)

        source_to_idx = {col: i for i, col in enumerate(source_columns)}
        for cm in config.columns:
            if cm.source not in source_to_idx:
                raise ValueError(
                    f"Source column '{cm.source}' not found in table "
                    f"'{config.source_table}'. Available: {source_columns}"
                )

        header_styles, column_widths, freeze_panes = self._extract_template_attrs(
            template_path, config.sheet, config.columns
        )

        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title=config.sheet)

        if freeze_panes:
            ws.freeze_panes = freeze_panes

        self._write_header(ws, config.columns, header_styles)

        for row in rows:
            data_row = [row[source_to_idx[cm.source]] for cm in config.columns]
            ws.append(data_row)

        wb.save(output_path)

        self._restore_column_widths(output_path, column_widths)

        context.logger.info(
            f"Output: wrote {len(rows)} rows to {output_path}"
        )

    def _extract_template_attrs(self, template_path, sheet_name, columns):
        wb = load_workbook(template_path)
        ws = wb[sheet_name] if sheet_name else wb.active

        first_row_cells = list(ws.iter_rows(min_row=1, max_row=1))
        if not first_row_cells or not first_row_cells[0]:
            wb.close()
            raise ValueError(
                f"Template sheet '{sheet_name}' has no data in row 1. "
                "The template must contain a header row."
            )
        template_headers = [c.value for c in first_row_cells[0]]
        # If the first row is completely empty, use column mapping as-is
        if all(h is None for h in template_headers):
            wb.close()
            return {}, None, None

        header_styles = {}
        for cell in first_row_cells[0]:
            if cell.value is not None:
                header_styles[str(cell.value)] = {
                    "font": copy.copy(cell.font),
                    "fill": copy.copy(cell.fill),
                    "border": copy.copy(cell.border),
                    "alignment": copy.copy(cell.alignment),
                    "number_format": cell.number_format,
                }

        column_widths = {}
        for col_letter, col_dim in ws.column_dimensions.items():
            if col_dim.width is not None:
                column_widths[col_letter] = col_dim.width

        for cm in columns:
            if cm.target not in header_styles:
                wb.close()
                raise ValueError(
                    f"Target column '{cm.target}' not found in template headers. "
                    f"Available: {template_headers}"
                )

        freeze_panes = ws.freeze_panes

        wb.close()
        return header_styles, column_widths, freeze_panes

    def _write_header(self, ws, columns, header_styles):
        header_row = []
        for cm in columns:
            cell = WriteOnlyCell(ws, value=cm.target)
            style = header_styles.get(cm.target)
            if style:
                cell.font = style["font"]
                cell.fill = style["fill"]
                cell.border = style["border"]
                cell.alignment = style["alignment"]
                cell.number_format = style["number_format"]
            header_row.append(cell)
        ws.append(header_row)

    def _restore_column_widths(self, output_path, column_widths):
        if not column_widths:
            return
        wb = load_workbook(output_path)
        ws = wb.active
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        wb.save(output_path)
        wb.close()
