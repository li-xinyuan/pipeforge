import tempfile
import os
import copy
import csv as csv_module
import sys

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


@pytest.fixture(autouse=True)
def _ensure_plugins_registered():
    """确保插件注册表在每项测试前处于完整状态。

    tests/test_registry.py 的 setup_method 调用 PluginRegistry.clear() 后无 teardown，
    导致后续测试（字母序在 r 之后，如 test_xml_input.py）找不到插件。
    configforge/tests/test_plugins_api.py 的 reload 也会改变类对象身份，使 ``is`` 检查失效。

    本 fixture 在每项测试前重新加载所有插件模块，触发 @register_plugin 重新注册。
    注意：reload 会创建新的类对象，测试中不应使用 ``is`` 检查插件类身份，
    应改用 ``cls.config_model() is XxxConfig``（config 类来自 config.models，不被 reload）。
    """
    import importlib
    import pkgutil

    from pipeforge.plugins._loader import load_all_plugins

    load_all_plugins()  # 确保所有插件模块已导入
    for subpkg in ("input", "processor", "output"):
        pkg = importlib.import_module(f"pipeforge.plugins.{subpkg}")
        for _finder, name, _ispkg in pkgutil.iter_modules(pkg.__path__):
            if name.startswith("_"):
                continue
            mod_path = f"pipeforge.plugins.{subpkg}.{name}"
            if mod_path in sys.modules:
                importlib.reload(sys.modules[mod_path])
    yield


@pytest.fixture
def sample_xlsx():
    """Create a test Excel file with person details."""
    wb = Workbook()
    ws = wb.active
    ws.title = "人员列表"
    ws.append(["工号", "姓名", "部门", "岗位"])
    ws.append(["001", "张三", "技术部", "工程师"])
    ws.append(["002", "李四", "产品部", "产品经理"])
    ws.append(["003", "王五", "技术部", "高级工程师"])

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    yield path
    os.unlink(path)


@pytest.fixture
def sample_xlsx_attendance():
    """Create a test Excel file with attendance data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "考勤统计"
    ws.append(["工号", "出勤天数", "迟到次数"])
    ws.append(["001", "22", "1"])
    ws.append(["002", "20", "4"])
    ws.append(["003", "21", "0"])

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    yield path
    os.unlink(path)


@pytest.fixture
def template_xlsx():
    """Create a styled Excel template for output."""
    wb = Workbook()
    ws = wb.active
    ws.title = "报表"

    header_font = Font(name="微软雅黑", bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    headers = ["姓名", "所属部门", "岗位", "出勤天数", "迟到次数", "状态"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = copy.copy(header_font)
        cell.fill = copy.copy(header_fill)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 15
    ws.freeze_panes = "A2"

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    yield path
    os.unlink(path)


@pytest.fixture
def csv_person_file():
    """Create a temporary CSV file with person data."""
    tmp = tempfile.NamedTemporaryFile(
        mode="w", suffix=".csv", delete=False, encoding="utf-8", newline=""
    )
    writer = csv_module.writer(tmp)
    writer.writerow(["姓名", "部门", "工号"])
    writer.writerow(["张三", "技术部", "001"])
    writer.writerow(["李四", "市场部", "002"])
    tmp.close()
    yield tmp.name
    os.unlink(tmp.name)
