import copy
import os
import tempfile

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from pipeforge.config.models import ColumnMapping, ExcelOutputConfig
from pipeforge.core.context import Context
from pipeforge.core.sqlite import SQLiteManager
from pipeforge.plugins.output import ExcelOutputPlugin as ImportedPlugin
from pipeforge.plugins.output.excel import ExcelOutputPlugin, resolve_filename


@pytest.fixture
def template_xlsx():
    """Create a template Excel file with styles and column widths."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    header_font = Font(name="Arial", bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    headers = ["Name", "Department", "Status"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = copy.copy(header_font)
        cell.fill = copy.copy(header_fill)

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    ws.freeze_panes = "A2"

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    return path


@pytest.fixture
def setup_context_and_db(template_xlsx):
    db = SQLiteManager()
    db.create_table("report_data", ["name", "dept", "status"])
    with db.transaction():
        db.insert_row("report_data", ("Zhang San", "Engineering", "Normal"))
        db.insert_row("report_data", ("Li Si", "Product", "Warning"))

    output_dir = tempfile.mkdtemp()
    context = Context(
        db=db, params={}, yaml_dir=os.path.dirname(template_xlsx),
        scene_name="Monthly Report"
    )
    yield db, context, template_xlsx, output_dir
    db.close()
    db.remove()
    for f in os.listdir(output_dir):
        os.unlink(os.path.join(output_dir, f))
    os.rmdir(output_dir)


class TestResolveFilename:
    def test_default_filename(self):
        result = resolve_filename(None, "Monthly Report")
        assert result.endswith(".xlsx")
        assert "Monthly Report" in result

    def test_scene_name_variable(self):
        result = resolve_filename("{{scene_name}}_report.xlsx", "Monthly Report")
        assert result == "Monthly Report_report.xlsx"

    def test_date_variable(self):
        result = resolve_filename("report_{{date:%Y}}.xlsx", "test")
        assert result.startswith("report_20")
        assert result.endswith(".xlsx")


class TestExcelOutputPlugin:
    def test_config_model(self):
        assert ExcelOutputPlugin.config_model() == ExcelOutputConfig

    def test_execute_writes_output(self, setup_context_and_db):
        db, context, template_xlsx, output_dir = setup_context_and_db
        config = ExcelOutputConfig(
            template=template_xlsx,
            sheet="Report",
            output_dir=output_dir,
            source_table="report_data",
            columns=[
                ColumnMapping(source="name", target="Name"),
                ColumnMapping(source="dept", target="Department"),
                ColumnMapping(source="status", target="Status"),
            ],
        )

        plugin = ExcelOutputPlugin()
        plugin.name = "excel"
        plugin.execute(context, config)

        assert context.output_path is not None
        assert os.path.exists(context.output_path)


class TestImportedPlugin:
    def test_plugin_importable_from_package(self):
        assert ImportedPlugin is ExcelOutputPlugin
