import io
import openpyxl


def read_excel_info(file_like, sheet_name=None, max_sample_rows=10):
    wb = openpyxl.load_workbook(file_like, read_only=True)
    try:
        sheets = wb.sheetnames
        ws_name = sheet_name or sheets[0]
        ws = wb[ws_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(next(rows_iter))]
        except StopIteration:
            return {"sheets": sheets, "columns": [], "sample_rows": []}
        sample_rows = []
        for i, row in enumerate(rows_iter):
            if i >= max_sample_rows:
                break
            sample_rows.append([str(v) if v is not None else "" for v in row])
        return {"sheets": sheets, "columns": headers, "sample_rows": sample_rows}
    finally:
        wb.close()
